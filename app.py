import os
import io
import base64
import time
import threading
from collections import defaultdict, deque

import requests
from flask import Flask, request, abort

from linebot import LineBotApi, WebhookHandler
from linebot.exceptions import InvalidSignatureError
from linebot.models import (
    MessageEvent,
    TextMessage,
    TextSendMessage,
    ImageMessage,
    FileMessage,
)

app = Flask(__name__)

# --- LINE credentials (from Render env vars) ---
LINE_CHANNEL_ACCESS_TOKEN = os.environ["LINE_CHANNEL_ACCESS_TOKEN"]
LINE_CHANNEL_SECRET = os.environ["LINE_CHANNEL_SECRET"]
PERPLEXITY_API_KEY = os.environ["PERPLEXITY_API_KEY"]

line_bot_api = LineBotApi(LINE_CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(LINE_CHANNEL_SECRET)

# --- Bot identity ---
try:
    BOT_INFO = line_bot_api.get_bot_info()
    BOT_USER_ID = BOT_INFO.user_id
    BOT_DISPLAY_NAME = BOT_INFO.display_name
except Exception:
    BOT_USER_ID = None
    BOT_DISPLAY_NAME = "Fujisan Assistant"

PERPLEXITY_URL = "https://api.perplexity.ai/chat/completions"

SYSTEM_PROMPT = (
    "あなたは富士山ワイナリーのチームチャット用アシスタントです。 "
    "必ず日本語で返信してください。ユーザーが英語や他の言語で書いていても、返信は日本語のみで行ってください。 "
    "簡潔に答えてください（特に指示がない限り3〜6個の箇条書きで）。 "
    "最近のグループ会話がコンテキストとして与えられ、それに加えて分析対象（画像、ファイル、質問）が示されます。 "
    "チームが議論している内容に関連した返信になるよう、会話のコンテキストを活用してください。 "
    "ぶどう畑の写真の場合：樹勢、樹冠、葉、病害の兆候、結実状況についてコメントしてください。 "
    "ワインラベルやビジネス文書の場合：主要な内容を要約してください。 "
    "PDFやスプレッドシートの場合：構造的に要約し、締切、大きな数値、異常値を指摘してください。 "
    "事実関係の質問には、最新のウェブ情報を活用して回答してください。"
)

# --- In-memory chat history per group/room/user ---
# key = source_id (group_id, room_id, or user_id)
# value = deque of recent message snippets, oldest first
HISTORY_MAX = 30
chat_history = defaultdict(lambda: deque(maxlen=HISTORY_MAX))
history_lock = threading.Lock()

# Cache display names for users in groups
display_name_cache = {}


# ---------------- helpers ----------------

def source_key(event) -> str:
    src = event.source
    if src.type == "group":
        return f"group:{src.group_id}"
    if src.type == "room":
        return f"room:{src.room_id}"
    return f"user:{src.user_id}"


def get_display_name(event) -> str:
    src = event.source
    user_id = getattr(src, "user_id", None)
    if not user_id:
        return "someone"
    if user_id in display_name_cache:
        return display_name_cache[user_id]
    try:
        if src.type == "group":
            profile = line_bot_api.get_group_member_profile(src.group_id, user_id)
        elif src.type == "room":
            profile = line_bot_api.get_room_member_profile(src.room_id, user_id)
        else:
            profile = line_bot_api.get_profile(user_id)
        name = profile.display_name or "someone"
    except Exception:
        name = "someone"
    display_name_cache[user_id] = name
    return name


def log_message(event, snippet: str):
    """Append a short message snippet to this chat's history."""
    key = source_key(event)
    name = get_display_name(event)
    ts = time.strftime("%H:%M")
    with history_lock:
        chat_history[key].append(f"[{ts}] {name}: {snippet}")


def history_context(event) -> str:
    key = source_key(event)
    with history_lock:
        lines = list(chat_history[key])
    if not lines:
        return "(no recent conversation in this chat yet)"
    return "\n".join(lines)


def is_mentioned(event) -> bool:
    msg = event.message
    mention = getattr(msg, "mention", None)
    if mention and getattr(mention, "mentionees", None):
        for m in mention.mentionees:
            if BOT_USER_ID and getattr(m, "user_id", None) == BOT_USER_ID:
                return True
    return False


def strip_mention(text: str) -> str:
    parts = text.strip().split(" ", 1)
    if parts and parts[0].startswith("@"):
        return parts[1] if len(parts) > 1 else ""
    return text


def download_content(message_id: str) -> bytes:
    content = line_bot_api.get_message_content(message_id)
    return b"".join(chunk for chunk in content.iter_content())


def safe_reply(event, text: str):
    line_bot_api.reply_message(event.reply_token, TextSendMessage(text=text[:4900]))


# ---------------- Perplexity ----------------

def ask_perplexity_text(user_text: str, model: str = "sonar") -> str:
    headers = {
        "Authorization": f"Bearer {PERPLEXITY_API_KEY}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": user_text},
        ],
    }
    try:
        r = requests.post(PERPLEXITY_URL, headers=headers, json=payload, timeout=90)
        r.raise_for_status()
        return r.json()["choices"][0]["message"]["content"].strip()
    except Exception as e:
        return f"(error contacting Perplexity: {e})"


def ask_perplexity_image(image_bytes: bytes, prompt: str) -> str:
    headers = {
        "Authorization": f"Bearer {PERPLEXITY_API_KEY}",
        "Content-Type": "application/json",
    }
    b64 = base64.b64encode(image_bytes).decode("utf-8")
    data_url = f"data:image/jpeg;base64,{b64}"
    payload = {
        "model": "sonar-pro",
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": data_url}},
                ],
            },
        ],
    }
    try:
        r = requests.post(PERPLEXITY_URL, headers=headers, json=payload, timeout=120)
        r.raise_for_status()
        return r.json()["choices"][0]["message"]["content"].strip()
    except Exception as e:
        return f"(error analyzing image: {e})"


# ---------------- file parsers ----------------

def extract_pdf_text(data: bytes, max_chars: int = 12000) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(data))
        chunks = [p.extract_text() or "" for p in reader.pages]
        text = "\n".join(chunks).strip()
        return text[:max_chars] if text else "(no extractable text in PDF)"
    except Exception as e:
        return f"(PDF parse error: {e})"


def extract_xlsx_summary(data: bytes, max_chars: int = 8000) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        out = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            out.append(f"## Sheet: {sheet_name}")
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    out.append(" | ".join(cells))
                    row_count += 1
                if row_count >= 60:
                    out.append("... (truncated)")
                    break
        joined = "\n".join(out)
        return joined[:max_chars] if joined else "(empty workbook)"
    except Exception as e:
        return f"(Excel parse error: {e})"


def extract_csv_summary(data: bytes, max_chars: int = 8000) -> str:
    try:
        text = data.decode("utf-8", errors="replace")
        lines = text.splitlines()[:60]
        if len(text.splitlines()) > 60:
            lines.append("... (truncated)")
        return "\n".join(lines)[:max_chars]
    except Exception as e:
        return f"(CSV parse error: {e})"


# ---------------- routes ----------------

@app.route("/callback", methods=["POST"])
def callback():
    signature = request.headers.get("X-Line-Signature", "")
    body = request.get_data(as_text=True)
    print("WEBHOOK BODY:", body, flush=True) 
    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        abort(400)
    return "OK"


@app.route("/", methods=["GET"])
def health():
    return f"Fujisan LINE bot running. Bot: {BOT_DISPLAY_NAME}"


# ---------------- event handlers ----------------

@handler.add(MessageEvent, message=TextMessage)
def handle_text(event):
    text = event.message.text or ""
    # Always log to history
    log_message(event, text)

    source_type = event.source.type

    # In 1:1 chats, always respond. In groups, only when mentioned.
    if source_type != "user" and not is_mentioned(event):
        return

    question = strip_mention(text).strip()
    if not question:
        question = "最近の会話について役に立つコメントをお願いします。"

    prompt = (
        "最近のグループ会話（古い順）:\n"
        f"{history_context(event)}\n\n"
        "----\n"
        f"最新のメッセージ: {question}\n\n"
        "最近の会話を踏まえて、日本語で丁寧に返信してください。"
    )
    reply = ask_perplexity_text(prompt, model="sonar")
    log_message(event, f"{BOT_DISPLAY_NAME}: {reply[:200]}")
    safe_reply(event, reply)


@handler.add(MessageEvent, message=ImageMessage)
def handle_image(event):
    log_message(event, "[sent a photo]")
    try:
        image_bytes = download_content(event.message.id)
    except Exception as e:
        safe_reply(event, f"(could not download image: {e})")
        return

    prompt = (
        "最近のグループ会話（古い順）:\n"
        f"{history_context(event)}\n\n"
        "----\n"
        "チームメンバーが添付の写真を共有しました。 "
        "写真を分析し（ぶどうの樹／樹冠／葉／果実／病害／ラベル等）、関連があれば最近の会話と結び付けてコメントしてください。 "
        "返信は必ず日本語で行ってください。"
    )
    reply = ask_perplexity_image(image_bytes, prompt)
    log_message(event, f"{BOT_DISPLAY_NAME}: {reply[:200]}")
    safe_reply(event, reply)


@handler.add(MessageEvent, message=FileMessage)
def handle_file(event):
    file_name = event.message.file_name or ""
    log_message(event, f"[sent file: {file_name}]")
    fn = file_name.lower()

    try:
        data = download_content(event.message.id)
    except Exception as e:
        safe_reply(event, f"(could not download file: {e})")
        return

    if fn.endswith(".pdf"):
        body = extract_pdf_text(data)
        kind = "PDF"
    elif fn.endswith((".xlsx", ".xls")):
        body = extract_xlsx_summary(data)
        kind = "spreadsheet"
    elif fn.endswith(".csv"):
        body = extract_csv_summary(data)
        kind = "CSV"
    else:
        return  # unsupported file type, stay silent

    prompt = (
        "最近のグループ会話（古い順）:\n"
        f"{history_context(event)}\n\n"
        "----\n"
        f"チームメンバーが「{file_name}」という{kind}を共有しました。 "
        "主要な内容を4〜8個の箇条書きで日本語で要約し、 "
        "注目すべき点（締切、大きな数値、異常値など）を指摘してください。 "
        "関連があれば最近の会話と結び付けてください。返信は必ず日本語で行ってください。\n\n"
        f"ファイル内容:\n{body}"
    )
    reply = ask_perplexity_text(prompt, model="sonar-pro")
    log_message(event, f"{BOT_DISPLAY_NAME}: {reply[:200]}")
    safe_reply(event, reply)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
